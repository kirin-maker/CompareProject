import json
import tkinter as tk
from tkinter import ttk, messagebox
import re
from deepdiff import DeepDiff
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
import os
from collections import OrderedDict

# ----------------- JSON Utility -----------------
def remove_description(obj):
    if isinstance(obj, dict):
        obj.pop("description", None)
        for v in obj.values():
            remove_description(v)
    elif isinstance(obj, list):
        for item in obj:
            remove_description(item)

def filter_out_debug(obj):
    if isinstance(obj, dict):
        return {k: filter_out_debug(v) for k, v in obj.items() if "debug" not in k.lower()}
    elif isinstance(obj, list):
        return [filter_out_debug(i) for i in obj]
    else:
        return obj

def build_partial_json(base, diff_paths):
    partial = {}
    for path in diff_paths:
        keys = re.findall(r"\['([^]]+)'\]|\[(\d+)\]", path)
        keys = [k[0] if k[0] else int(k[1]) for k in keys]
        current_src = base
        current_partial = partial
        parents = [] # To keep track of parent objects/lists for updating references
        for i, key in enumerate(keys):
            is_last = (i == len(keys) - 1)
            
            # Navigate current_src
            if isinstance(current_src, dict):
                if key not in current_src:
                    break # Path not found in source
            elif isinstance(current_src, list):
                if not isinstance(key, int) or key >= len(current_src):
                    break # Path not found in source or invalid index
            else:
                break # Not a dict or list, cannot navigate further

            if isinstance(key, int): # Handling list items
                if not isinstance(current_partial, list):
                    if isinstance(current_partial, dict) and not current_partial:
                        new_list = []
                        if parents: # Update parent reference
                            parent, parent_key = parents[-1]
                            if isinstance(parent, dict):
                                parent[parent_key] = new_list
                            elif isinstance(parent, list):
                                parent[parent_key] = new_list
                        else: # This is the root level
                            partial = new_list
                        current_partial = new_list
                    else: # current_partial is not a list and not an empty dict, so cannot proceed
                        break
                
                # Ensure list is long enough for the index
                while len(current_partial) <= key:
                    current_partial.append({}) # Fill with empty dicts or appropriate default
                
                if is_last:
                    current_partial[key] = current_src[key]
                else:
                    parents.append((current_partial, key))
                    current_partial = current_partial[key]
                    current_src = current_src[key]

            else: # Handling dictionary keys (string key)
                if not isinstance(current_partial, dict):
                    break # Not a dict, cannot add string key
                
                if key not in current_partial:
                    current_partial[key] = {} # Create empty dict for nested structure
                
                if is_last:
                    current_partial[key] = current_src[key]
                else:
                    parents.append((current_partial, key))
                    current_partial = current_partial[key]
                    current_src = current_src[key]
    return partial

def format_full_output(data):
    if not isinstance(data, dict):
        return json.dumps(data, indent=2, ensure_ascii=False)
    output_lines = []
    
    # Process promoInfo first if it exists
    if "promoInfo" in data and isinstance(data["promoInfo"], list):
        sorted_promos = sorted(
            data["promoInfo"],
            key=lambda p: int(p.get("promoNumber", "0")) if str(p.get("promoNumber", "0")).isdigit() else float('inf')
        )
        for promo in sorted_promos:
            promo_number = promo.get("promoNumber", "N/A")
            output_lines.append(f"promoNumber: {promo_number}")
            output_lines.append(json.dumps(promo, indent=2, ensure_ascii=False))
            output_lines.append("")
    
    other_keys = [k for k in data.keys() if k != "promoInfo"]

    for key in other_keys:
        value = data[key]
        output_lines.append(f'"{key}": {json.dumps(value, indent=2, ensure_ascii=False)}')
        output_lines.append("")
    
    return "\n".join(output_lines).strip()

# ----------------- GUI Utility -----------------
def clear_label_result():
    label_result.config(text="")

def copy_text(widget):
    content = widget.get("1.0", tk.END).strip()
    if content:
        try:
            root.clipboard_clear()
            root.clipboard_append(content)
            label_result.config(text="✅ Text copied to clipboard!", foreground="#66ff99")
            root.after(1500, clear_label_result)
        except Exception as e:
            label_result.config(text=f"❌ Copy failed: {e}", foreground="#ff6666")
            root.after(1500, clear_label_result)
    else:
        label_result.config(text="⚠️ Nothing to copy.", foreground="#ffaa00")
        root.after(1500, clear_label_result)

def add_right_click_menu(widget):
    menu = tk.Menu(widget, tearoff=0, bg="#2e2e2e", fg="#f8f8f2")
    menu.add_command(label="Paste", command=lambda: widget.event_generate("<<Paste>>"))
    widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

def bind_scroll(widget):
    widget.bind("<Enter>", lambda e: widget.bind_all("<MouseWheel>", lambda ev: widget.yview_scroll(int(-1*(ev.delta/120)), "units")))
    widget.bind("<Leave>", lambda e: widget.unbind_all("<MouseWheel>"))

def bind_paste_shortcuts(widget):
    def do_paste(event):
        widget.event_generate("<<Paste>>")
        return "break"
    for seq in ("<Control-v>", "<Control-V>", "<Shift-Insert>"):
        widget.bind(seq, do_paste)

def highlight_promo_lines(text_widget):
    text_widget.tag_configure("highlight", foreground="#00ff00", font=("Segoe UI", 10, "bold"))
    start = "1.0"
    while True:
        pos = text_widget.search(r"^=+ promoNumber: .* =+$", start, stopindex=tk.END, regexp=True)
        if not pos:
            break
        text_widget.tag_add("highlight", pos, f"{pos} lineend")
        start = f"{pos} lineend"

def highlight_differences(text_widget, diff_paths):
    text_widget.tag_remove("diff_highlight", "1.0", tk.END)
    text_widget.tag_configure("diff_highlight", foreground="#F700FF", font=("Segoe UI", 10, "bold"))
    for path in diff_paths:
        
        # Example: path = "['promoInfo'][0]['name']" -> last_key = "name"
        keys_in_path = re.findall(r"\['([^]]+)'\]|\[(\d+)\]", path)
        if not keys_in_path:
            continue
        
        # Extract the relevant key or index to search for
        last_key_or_index = keys_in_path[-1][0] or keys_in_path[-1][1]
        
        # Simple search for the last key/value on a line
        search_pattern = re.escape(f'"{last_key_or_index}"') if not str(last_key_or_index).isdigit() else re.escape(str(last_key_or_index))
        
        start_pos = "1.0"
        while True:
            # Search for the key on its own line or within a line
            pos = text_widget.search(search_pattern, start_pos, stopindex=tk.END, regexp=True)
            if not pos:
                break
            
            # Highlight the entire line where the key is found
            line_start = f"{pos.split('.')[0]}.0"
            line_end = f"{pos.split('.')[0]}.end"
            text_widget.tag_add("diff_highlight", line_start, line_end)
            
            # Continue searching from the end of the current highlighted line
            start_pos = f"{pos} lineend"

# ----------------- Global Variables -----------------
EXPORT_FOLDER = os.path.join(os.getcwd(), "export")
EXCEL_PATH = os.path.join(EXPORT_FOLDER, "Compare_Export.xlsx")

# Create export folder if it doesn't exist
if not os.path.exists(EXPORT_FOLDER):
    try:
        os.makedirs(EXPORT_FOLDER)
    except Exception as e:
        messagebox.showerror("Folder Creation Failed", f"Could not create the export folder:\n{e}")
        raise

# last_export_data will store (partial_base_result, partial_compare_result, total_diff_paths)
last_export_data = None 

# ----------------- Excel Export Utility -----------------

def filter_out_debug(data):
    if isinstance(data, dict):
        keys_to_remove = ["debug", "qualifySpend", "quantity", "numberOfTotalSavers"]
        for key in keys_to_remove:
            data.pop(key, None)
        for v in data.values():
            filter_out_debug(v)
    elif isinstance(data, list):
        for item in data:
            filter_out_debug(item)
    return data

def to_pretty_json_blocks(promo_list):
    blocks = []
    for promo in promo_list:
        if not isinstance(promo, dict):
            try:
                promo = json.loads(str(promo))
            except Exception:
                continue

        # ดึงและแปลง promoNumber
        promo_number_raw = promo.get("promoNumber", "UNKNOWN")
        try:
            promo_number = int(promo_number_raw)
        except Exception:
            promo_number = promo_number_raw  # fallback เช่น "UNKNOWN"

        header = f"{promo_number}"  # ใช้เลขอย่างเดียว
        pretty_json = json.dumps(promo, indent=2, ensure_ascii=False)
        blocks.append(f"{header}\n{pretty_json}")
    return blocks


def filter_out_debug(data):
    if isinstance(data, dict):
        keys_to_remove = ["debug", "qualifySpend", "quantity", "numberOfTotalSavers"]
        for key in keys_to_remove:
            data.pop(key, None)
        for v in data.values():
            filter_out_debug(v)
    elif isinstance(data, list):
        for item in data:
            filter_out_debug(item)
    return data

def to_pretty_json_blocks(promo_list):
    blocks = []
    for promo in promo_list:
        if not isinstance(promo, dict):
            try:
                promo = json.loads(str(promo))
            except Exception:
                continue

        # ดึงและแปลง promoNumber
        promo_number_raw = promo.get("promoNumber", "UNKNOWN")
        try:
            promo_number = int(promo_number_raw)
        except Exception:
            promo_number = promo_number_raw  # fallback เช่น "UNKNOWN"

        header = f"{promo_number}"  # ใช้เลขอย่างเดียว
        pretty_json = json.dumps(promo, indent=2, ensure_ascii=False)
        blocks.append(f"{header}\n{pretty_json}")
    return blocks


def write_lines_aligned_to_excel(ws, start_row, base_lines, compare_lines, diff_fill, align_top_wrap):
    row = start_row

    len_b, len_c = len(base_lines), len(compare_lines)
    i, j = 0, 0

    def extract_key(line):
        stripped = line.lstrip()
        if ":" in stripped:
            return stripped.split(":", 1)[0].strip().strip('"')
        return None

    while i < len_b or j < len_c:
        b_line = base_lines[i] if i < len_b else None
        c_line = compare_lines[j] if j < len_c else None

        b_key = extract_key(b_line) if b_line is not None else None
        c_key = extract_key(c_line) if c_line is not None else None

        # กรณี key ตรงกันหรือทั้งสอง None (บรรทัดปกติ)
        if b_key == c_key:
            val_b = b_line if b_line is not None else ""
            val_c = c_line if c_line is not None else ""
            i += 1
            j += 1

        # กรณี key ต่างกัน, แต่ key c_key มีใน base_lines ข้างหน้า
        elif c_key is not None and (b_key != c_key):
            found_idx = None
            for k in range(i + 1, len_b):
                if extract_key(base_lines[k]) == c_key:
                    found_idx = k
                    break
            if found_idx is not None:
                # เติม "Nodata" ใน compare_lines จนกว่าจะเจอ key c_key ใน base_lines
                val_b = b_line if b_line is not None else ""
                val_c = ""
                i += 1
            else:
                # กรณี key c_key ไม่มีใน base_lines, เติม "Nodata" ใน base_lines
                val_b = ""
                val_c = c_line if c_line is not None else ""
                j += 1

        # กรณี key ต่างกัน, แต่ key b_key มีใน compare_lines ข้างหน้า
        elif b_key is not None and (c_key != b_key):
            found_idx = None
            for k in range(j + 1, len_c):
                if extract_key(compare_lines[k]) == b_key:
                    found_idx = k
                    break
            if found_idx is not None:
                # เติม "Nodata" ใน base_lines จนกว่าจะเจอ key b_key ใน compare_lines
                val_b = ""
                val_c = c_line if c_line is not None else ""
                j += 1
            else:
                # กรณี key b_key ไม่มีใน compare_lines, เติม "Nodata" ใน compare_lines
                val_b = b_line if b_line is not None else ""
                val_c = ""
                i += 1

        else:
            # กรณีอื่น ๆ เติม "Nodata" หากไม่มีบรรทัด
            val_b = b_line if b_line is not None else ""
            val_c = c_line if c_line is not None else ""
            i += (i < len_b)
            j += (j < len_c)

        cell_b = ws.cell(row=row, column=2, value=val_b)
        cell_c = ws.cell(row=row, column=3, value=val_c)

        cell_b.alignment = align_top_wrap
        cell_c.alignment = align_top_wrap

        # ไฮไลท์ถ้าข้อความต่างกัน (เว้นว่าง "" ไม่ไฮไลท์)
        if val_b.strip() != val_c.strip():
            if val_b.strip() != "":
                cell_b.fill = diff_fill
            if val_c.strip() != "":
                cell_c.fill = diff_fill

        row += 1


def export_to_excel():
    if not last_export_data or len(last_export_data) != 2:
        messagebox.showwarning("No Comparison Data", "Please compare JSON files before exporting.")
        return

    base_text, compare_text = last_export_data
    base_lines = base_text.splitlines()
    compare_lines = compare_text.splitlines()

    diff_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    align_top_wrap = Alignment(vertical="top", wrap_text=True)

    try:
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
        else:
            wb = Workbook()
            wb.remove(wb.active)
    except Exception as e:
        messagebox.showerror("Excel Load Error", str(e))
        return

    # ลบชีทเดิมหากมี และสร้างชีทใหม่
    if "Comparison" in wb.sheetnames:
        del wb["Comparison"]
    ws = wb.create_sheet("Comparison")

    # ปรับความกว้างคอลัมน์
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80

    # ปรับความสูงแถว
    ws.row_dimensions[2].height = 140

    # ✅ ดึงข้อมูลจาก GUI
    try:
        raw_request = text_request.get("1.0", tk.END).strip()  # ข้อความดิบ (ไม่แปลง JSON)
        raw_newpro = text_base.get("1.0", tk.END).strip()
        raw_online = text_compare.get("1.0", tk.END).strip()

        # พยายามแปลง JSON เฉพาะ Newpro
        try:
            newpro_obj = json.loads(raw_newpro) if raw_newpro else {}
            res_newpro_text = json.dumps(newpro_obj, ensure_ascii=False, indent=2)
        except Exception:
            res_newpro_text = raw_newpro  # ถ้าแปลงไม่ได้ ใช้ข้อความดิบ

        # พยายามแปลง JSON เฉพาะ Online
        try:
            online_obj = json.loads(raw_online) if raw_online else {}
            res_online_text = json.dumps(online_obj, ensure_ascii=False, indent=2)
        except Exception:
            res_online_text = raw_online  # ถ้าแปลงไม่ได้ ใช้ข้อความดิบ

    except Exception as e:
        messagebox.showerror("Input Error", f"Unable to read inputs: {e}")
        return

    # ✅ เขียน Header
    ws.cell(row=1, column=1, value="Request_Promotion")
    ws.cell(row=1, column=2, value="Newproengine_Response")
    ws.cell(row=1, column=3, value="LP_Response")

    # ✅ เขียนข้อมูล JSON input และข้อความดิบแบบ 1 บรรทัด พร้อมจัดรูปแบบ alignment
    input_align = Alignment(vertical="top", horizontal="left", wrap_text=True)

    cell_req = ws.cell(row=2, column=1, value=raw_request)
    cell_req.alignment = input_align

    cell_newpro = ws.cell(row=2, column=2, value=res_newpro_text)
    cell_newpro.alignment = input_align

    cell_online = ws.cell(row=2, column=3, value=res_online_text)
    cell_online.alignment = input_align

    # เว้นบรรทัดก่อนเริ่มเปรียบเทียบ
    ws.cell(row=3, column=1, value="")

    # เขียน Header เปรียบเทียบ
    ws.cell(row=4, column=2, value="Newproengine_Diffrent")
    ws.cell(row=4, column=3, value="LP_Diffrent")

    # เขียนข้อมูลเปรียบเทียบจาก GUI ต่อจากแถวที่ 5
    write_lines_aligned_to_excel(ws, 5, base_lines, compare_lines, diff_fill, align_top_wrap)

    try:
        wb.save(EXCEL_PATH)
        messagebox.showinfo("Export Successful", f"Excel file saved to:\n{EXCEL_PATH}")
    except PermissionError:
        messagebox.showerror("Save Failed", "Permission denied. Please close the Excel file and try again.")
    except Exception as e:
        messagebox.showerror("Save Failed", f"An unexpected error occurred:\n{e}")


# ----------------- Core Function: compare_json ----------------- #===================อย่าแก้ไขส่วนนี้ลงไป===================


def compare_json():
    try:
        base_data = json.loads(text_base.get("1.0", tk.END))
        compare_data = json.loads(text_compare.get("1.0", tk.END))
    except json.JSONDecodeError as e:
        messagebox.showerror("รูปแบบ JSON ไม่ถูกต้อง", str(e))
        return

    remove_description(base_data)
    remove_description(compare_data)

    base_filtered = filter_out_debug(base_data)
    compare_filtered = filter_out_debug(compare_data)

    base_promos = {p["promoNumber"]: p for p in base_filtered.get("promoInfo", []) if "promoNumber" in p}
    compare_promos = {p["promoNumber"]: p for p in compare_filtered.get("promoInfo", []) if "promoNumber" in p}

    partial_base_result = {"promoInfo": []}
    partial_compare_result = {"promoInfo": []}
    total_diff_paths = []

    all_promo_numbers = sorted(set(base_promos.keys()) | set(compare_promos.keys()), key=lambda x: int(x))

    for promo_num in all_promo_numbers:
        base_promo = base_promos.get(promo_num)
        compare_promo = compare_promos.get(promo_num)

        if base_promo and compare_promo:
            # กรณีมีทั้งสองฝั่ง
            diff = DeepDiff(base_promo, compare_promo, ignore_order=False, report_repetition=True, view="tree")

            if not diff:
                continue  # ไม่มี diff ก็ไม่ต้องใส่

            path_list = []
            for section in diff:
                for change in diff[section]:
                    if hasattr(change, 'path'):
                        path = change.path(output_format='list')
                        s = "".join(f"[{p}]" if isinstance(p, int) else f"['{p}']" for p in path)
                        path_list.append(s)

            total_diff_paths.extend([f"['promoInfo'][{len(partial_base_result['promoInfo'])}]{p}" for p in path_list])

            partial_base = build_partial_json(base_promo, path_list)
            partial_base["promoNumber"] = promo_num
            partial_compare = build_partial_json(compare_promo, path_list)
            partial_compare["promoNumber"] = promo_num

        elif base_promo and not compare_promo:
            # มีเฉพาะใน Base
            partial_base = base_promo.copy()
            partial_base["promoNumber"] = promo_num
            partial_compare = {"promoNumber": promo_num}  # ว่างเปล่า

        elif compare_promo and not base_promo:
            # มีเฉพาะใน Compare
            partial_compare = compare_promo.copy()
            partial_compare["promoNumber"] = promo_num
            partial_base = {"promoNumber": promo_num}  # ว่างเปล่า

        partial_base_result["promoInfo"].append(partial_base)
        partial_compare_result["promoInfo"].append(partial_compare)

    # ==== เปรียบเทียบฟิลด์อื่น ๆ ที่ไม่ใช่ promoInfo ====
    other_keys = set(base_filtered.keys()) | set(compare_filtered.keys())
    other_keys.discard("promoInfo")

    for key in sorted(other_keys):
        if key not in base_filtered or key not in compare_filtered:
            continue

        diff = DeepDiff(base_filtered[key], compare_filtered[key], ignore_order=False, report_repetition=True, view="tree")

        if not diff:
            continue

        path_list = []
        for section in diff:
            for change in diff[section]:
                if hasattr(change, 'path'):
                    path = change.path(output_format='list')
                    s = f"['{key}']" + "".join(f"[{p}]" if isinstance(p, int) else f"['{p}']" for p in path)
                    path_list.append(s)

        total_diff_paths.extend(path_list)

        partial_base = build_partial_json(base_filtered, path_list)
        partial_compare = build_partial_json(compare_filtered, path_list)

        partial_base_result.update(partial_base)
        partial_compare_result.update(partial_compare)

    # ==== สร้างผลลัพธ์และแสดงผล ====
    base_result = format_full_output(partial_base_result)
    compare_result = format_full_output(partial_compare_result)

    text_partial_base.delete("1.0", tk.END)
    text_partial_compare.delete("1.0", tk.END)
    text_partial_base.insert(tk.END, base_result)
    text_partial_compare.insert(tk.END, compare_result)

    highlight_promo_lines(text_partial_base)
    highlight_promo_lines(text_partial_compare)
    highlight_differences(text_partial_base, total_diff_paths)
    highlight_differences(text_partial_compare, total_diff_paths)

    global last_export_data
    last_export_data = (
        text_partial_base.get("1.0", tk.END).strip(),
        text_partial_compare.get("1.0", tk.END).strip()
    )

    label_result.config(text=f"🔍 พบความแตกต่างทั้งหมด {len(total_diff_paths)} จุด")
# ----------------- GUI Setup -----------------
root = tk.Tk()
root.title("🧠 JSON Compare Tool")
root.attributes("-fullscreen", True)

is_fullscreen = True
def toggle_fullscreen(event=None):
    global is_fullscreen
    is_fullscreen = not is_fullscreen
    root.attributes("-fullscreen", is_fullscreen)
def exit_fullscreen(event=None):
    root.attributes("-fullscreen", False)
root.bind("<F11>", toggle_fullscreen)
root.bind("<Escape>", exit_fullscreen)

DARK_BG = "#2e2e2e"
DARK_TEXT = "#f8f8f2"
TEXTBOX_BG = "#1e1e1e"
HIGHLIGHT = "#3c3f41"

root.configure(bg=DARK_BG)
style = ttk.Style()
style.theme_use("clam")
style.configure("TFrame", background=DARK_BG)
style.configure("TLabel", background=DARK_BG, foreground=DARK_TEXT)
style.configure("Header.TLabel", font=("Segoe UI", 13, "bold"), background=DARK_BG, foreground=DARK_TEXT)
style.configure("TButton", background=HIGHLIGHT, foreground="#ffffff", relief="flat", padding=6)
style.map("TButton", background=[("active", "#505354")], foreground=[("active", "#ffffff")])
style.configure("TLabelframe", background=DARK_BG, foreground=DARK_TEXT)
style.configure("TLabelframe.Label", background=DARK_BG, foreground=DARK_TEXT)

# กำหนดน้ำหนักคอลัมน์และแถวใหม่
root.grid_columnconfigure(0, weight=1)  # ซ้ายสุด (Request_Promotion)
root.grid_columnconfigure(1, weight=3)  # ขวา (LP, Pro Engine, Controls, Output)
root.grid_rowconfigure(5, weight=1)     # แถวล่างสุด (Output)

top_frame = ttk.Frame(root)
top_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(10, 5))
top_frame.columnconfigure(0, weight=1)
ttk.Label(top_frame, text="🧠 JSON Compare Tool", style="Header.TLabel").pack()

# --- แยก Frame สำหรับ Request_Promotion (ฝั่งซ้าย) ---
frame_request = ttk.Frame(root)
frame_request.grid(row=1, column=0, rowspan=5, sticky="nsew", padx=(10, 5), pady=10)
frame_request.grid_rowconfigure(1, weight=1)

ttk.Label(frame_request, text="📝 Request_Promotion", style="Header.TLabel").grid(row=0, column=0, sticky="w")
text_request = tk.Text(frame_request, bg=TEXTBOX_BG, fg=DARK_TEXT, insertbackground="white", relief="groove")
text_request.grid(row=1, column=0, sticky="nsew")
add_right_click_menu(text_request)
bind_scroll(text_request)
bind_paste_shortcuts(text_request)

# --- Frame หลักฝั่งขวา (LP, Pro Engine, Controls, Output) ---
frame_input = ttk.Frame(root)
frame_input.grid(row=1, column=1, sticky="nsew", padx=10)
frame_input.grid_columnconfigure(0, weight=1)  # LP
frame_input.grid_columnconfigure(1, weight=1)  # Pro Engine
frame_input.grid_rowconfigure(1, weight=1)     # ขยายความสูง

# 📘 LP
ttk.Label(frame_input, text="📘 LP", style="Header.TLabel").grid(row=0, column=0, sticky="w")
text_base = tk.Text(frame_input, bg=TEXTBOX_BG, fg=DARK_TEXT, insertbackground="white", relief="groove", height=18)
text_base.grid(row=2, column=0, sticky="nsew", padx=(0,5))
add_right_click_menu(text_base)
bind_scroll(text_base)
bind_paste_shortcuts(text_base)

# 📙 Pro Engine
ttk.Label(frame_input, text="📙 Pro Engine", style="Header.TLabel").grid(row=0, column=1, sticky="w")
text_compare = tk.Text(frame_input, bg=TEXTBOX_BG, fg=DARK_TEXT, insertbackground="white", relief="groove", height=18)
text_compare.grid(row=2, column=1, sticky="nsew", padx=(5,0))
add_right_click_menu(text_compare)
bind_scroll(text_compare)
bind_paste_shortcuts(text_compare)

ttk.Button(root, text="🔍 Compare JSON", command=compare_json).grid(row=2, column=1, pady=10)

label_result = ttk.Label(root, text="", background=DARK_BG, font=("Segoe UI", 12, "bold"))
label_result.grid(row=3, column=1, pady=5)

frame_controls = ttk.Frame(root)
frame_controls.grid(row=4, column=1, pady=5)
ttk.Button(frame_controls, text="📋 Copy Pro Engine Diff", command=lambda: copy_text(text_partial_base)).pack(side="left", padx=15)
ttk.Button(frame_controls, text="📤 Export to Excel", command=export_to_excel).pack(side="left", padx=15)
ttk.Button(frame_controls, text="📋 Copy LP Diff", command=lambda: copy_text(text_partial_compare)).pack(side="left", padx=15)

frame_output = ttk.Frame(root)
frame_output.grid(row=5, column=1, sticky="nsew", padx=10, pady=(0, 10))
frame_output.grid_columnconfigure(0, weight=1)
frame_output.grid_columnconfigure(1, weight=1)
frame_output.grid_rowconfigure(1, weight=1)

ttk.Label(frame_output, text="📘 LP Differences", style="Header.TLabel").grid(row=0, column=1, sticky="w")
text_partial_base = tk.Text(frame_output, bg=TEXTBOX_BG, fg=DARK_TEXT, insertbackground="white", relief="ridge")
text_partial_base.grid(row=1, column=0, sticky="nsew", padx=(0, 5))
add_right_click_menu(text_partial_base)
bind_scroll(text_partial_base)
bind_paste_shortcuts(text_partial_base)

ttk.Label(frame_output, text="📙 Pro Engine Differences", style="Header.TLabel").grid(row=0, column=0, sticky="w")
text_partial_compare = tk.Text(frame_output, bg=TEXTBOX_BG, fg=DARK_TEXT, insertbackground="white", relief="ridge")
text_partial_compare.grid(row=1, column=1, sticky="nsew", padx=(5, 0))
add_right_click_menu(text_partial_compare)
bind_scroll(text_partial_compare)
bind_paste_shortcuts(text_partial_compare)

root.mainloop()
